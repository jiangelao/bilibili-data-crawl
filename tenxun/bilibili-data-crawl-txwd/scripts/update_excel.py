#!/usr/bin/env python3
"""更新Excel表格，追加视频数据"""

import sys
import json
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, numbers
except ImportError:
    print(json.dumps({
        'success': False,
        'message': '缺少openpyxl依赖，请安装：pip install openpyxl'
    }, ensure_ascii=False))
    sys.exit(1)


# 严格按照 skill 规范定义的列顺序
HEADERS = [
    '统计日期',
    '播放量',
    '点赞量',
    '收藏量',
    '收播率(%)',
    '点赞率(%)',
    '封标点击率(星)',
    '3秒跳出率(%)',
    '互动率(%)',
    '播转粉率(%)',
    '游客占比(%)',
    '粉丝观看率(%)',
    '平均播放占总时长率(%)',
]

# 百分比列的 Excel 数字格式（保留2位小数，显示 % 符号）
PCT_FORMAT = '0.00"%"'


def sanitize_filename(filename: str) -> str:
    """清理文件名，移除非法字符"""
    illegal_chars = '<>:"/\\|?*'
    for char in illegal_chars:
        filename = filename.replace(char, '_')
    if len(filename) > 200:
        filename = filename[:200]
    return filename.strip()


def calculate_rates(data: dict) -> dict:
    """计算收播率和点赞率"""
    views = data.get('views', 0)
    likes = data.get('likes', 0)
    favorites = data.get('favorites', 0)

    result = {}
    if views > 0:
        result['favorite_rate'] = round(favorites / views * 100, 2)
        result['like_rate'] = round(likes / views * 100, 2)
    else:
        result['favorite_rate'] = 0.0
        result['like_rate'] = 0.0
    return result


def build_row(data: dict) -> list:
    """按 HEADERS 顺序构建数据行，百分比字段返回数值"""
    rates = calculate_rates(data)
    return [
        data.get('stat_date', datetime.now().strftime('%Y-%m-%d')),
        data.get('views', 0),
        data.get('likes', 0),
        data.get('favorites', 0),
        rates['favorite_rate'],
        rates['like_rate'],
        data.get('cover_click_rate', 0.0),
        data.get('bounce_rate_3s', 0.0),
        data.get('interaction_rate', 0.0),
        data.get('fan_conversion_rate', 0.0),
        data.get('tourist_ratio', 0.0),
        data.get('fan_view_rate', 0.0),
        data.get('avg_play_completion_rate', 0.0),
    ]


def apply_number_formats(ws, total_rows: int):
    """为百分比列应用 Excel 数字格式（跳过表头行）"""
    # 索引从 0 开始: 收播率=4, 点赞率=5, 3秒跳出率=7, 互动率=8,
    #                播转粉率=9, 游客占比=10, 粉丝观看率=11, 平均播放时长=12
    pct_cols = [4, 5, 7, 8, 9, 10, 11, 12]
    for col_idx in pct_cols:
        for row in range(2, total_rows + 1):
            cell = ws.cell(row=row, column=col_idx + 1)  # openpyxl 1-indexed
            if isinstance(cell.value, (int, float)):
                cell.number_format = PCT_FORMAT


def update_excel(title: str, data_json: str, save_dir: str) -> dict:
    """更新Excel表格

    Args:
        title: 视频标题
        data_json: JSON格式的视频数据
        save_dir: 用户工作区根目录（文件将保存到 {save_dir}/static_data/）

    Returns:
        dict: 操作结果
    """
    result = {
        'success': False,
        'file_path': '',
        'total_rows': 0,
        'message': ''
    }

    try:
        data = json.loads(data_json)
        row_data = build_row(data)

        # 保存目录固定为 {save_dir}/static_data/
        static_dir = Path(save_dir) / 'static_data'
        try:
            static_dir.mkdir(parents=True, exist_ok=True)
        except PermissionError:
            result['message'] = f'无法创建目录，权限不足：{static_dir}'
            return result
        except OSError as e:
            result['message'] = f'创建目录失败：{str(e)}'
            return result

        # 生成文件名
        safe_title = sanitize_filename(title)
        if not safe_title:
            safe_title = 'untitled'
        filename = f'{safe_title}.xlsx'
        file_path = static_dir / filename

        if file_path.exists():
            try:
                wb = load_workbook(file_path)
                ws = wb.active
                last_row = ws.max_row
                ws.append(row_data)
                total_rows = last_row + 1
                apply_number_formats(ws, total_rows)
                wb.save(file_path)
                result['success'] = True
                result['file_path'] = str(file_path)
                result['total_rows'] = total_rows
                result['message'] = f'数据已追加到现有文件，当前共 {total_rows} 行（含表头）'
            except Exception as e:
                result['message'] = f'追加数据失败：{str(e)}'
                return result
        else:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = '视频数据'

                # 写入表头
                ws.append(HEADERS)

                # 表头样式
                header_font = Font(bold=True)
                for cell in ws[1]:
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # 写入数据
                ws.append(row_data)

                # 设置数字格式
                total_rows = 2
                apply_number_formats(ws, total_rows)

                # 调整列宽
                column_widths = [12, 12, 10, 10, 12, 10, 15, 12, 12, 12, 12, 12, 20]
                for i, width in enumerate(column_widths, 1):
                    ws.column_dimensions[chr(64 + i)].width = width

                wb.save(file_path)

                result['success'] = True
                result['file_path'] = str(file_path)
                result['total_rows'] = 2
                result['message'] = '已创建新文件并写入数据'
            except Exception as e:
                result['message'] = f'创建文件失败：{str(e)}'
                return result

        return result

    except json.JSONDecodeError as e:
        result['message'] = f'JSON解析失败：{str(e)}'
        return result
    except Exception as e:
        result['message'] = f'未知错误：{str(e)}'
        return result


if __name__ == '__main__':
    if len(sys.argv) < 4:
        print(json.dumps({
            'success': False,
            'message': '参数不足，需要提供 视频标题、JSON数据、保存目录'
        }, ensure_ascii=False))
        sys.exit(1)

    title = sys.argv[1]
    data_json = sys.argv[2]
    save_dir = sys.argv[3]

    result = update_excel(title, data_json, save_dir)
    print(json.dumps(result, ensure_ascii=False))

    sys.exit(0 if result['success'] else 1)
